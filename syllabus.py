#
# シラバスのExcelファイルの変換
# 

import openpyxl

# ExcelからMapにあるセルを読んで，セルのリストを返す
def read_cells(in_sheet, in_map, offset):
    in_cells = []
    i = 0
    for pt in in_map :
        in_cells.append(in_sheet.cell(row=pt[0]+offset[0], column=pt[1]+offset[1]))
        i = i + 1
    return in_cells

# セルのリストをMapに従ってExcelに書き込む
def write_cells(out_sheet, in_cells, out_map, offset):
    i = 0
    for cell in in_cells :
        out_sheet.cell(row=out_map[i][0]+offset[0], column=out_map[i][1]+offset[1], value=cell.value)
        i = i + 1
    return

# 入力ファイルを出力ファイルにマップの指定に従って書き込む
def syllabus(in_filename, out_filename, in_map_list, out_map_list):
    # ワークブック
    in_wb = openpyxl.load_workbook(in_filename)
    # 出力ファイルは新規作成（テンプレートをベースに）
    out_wb = openpyxl.load_workbook('template.xlsx')

    # セルの読み書き
    for in_sheet in in_wb :
        # シート名が科目名
        kamoku = in_sheet.title
        # 講義回数を求め，それに応じたテンプレートとマップを選択
        # セル6E(6,5)（単位数）と「使用図書」の場所で講義回数を決定する
        # 単位数1,セル21A（）が「使用図書」である　→講義回数8（1学期での1単位科目）
        # 単位数1,セル21A（）が「使用図書」ではない→講義回数15（2学期での1単位科目）
        # 単位数2,セル21A（）が「使用図書」である→講義回数8（企業内実習I）
        # 単位数2,セル43A（）が「使用図書」である→講義回数30（ソーシャルデザイン実習）
        # 単位数2,セル21A（）が「使用図書」ではない→講義回数15（一般的な2単位科目）
        # 単位数5,セル25A（）が「使用図書」である→講義回数12（新規商品開発・販売実習Ⅰ, II）
        # 単位数5,のその他→講義回数14（臨地実務実習I）
        # 単位数15→講義回数26（臨地実務実習II）
        # 単位数4，セル28A(28,1)が「使用図書」である→講義回数15（企業設立実習I，II，III, 事業計画策定総合実習）
        # 単位数4，セル28A(25,1)が「使用図書」である→講義回数12（企業内実習II）
        # 単位数4，セル28A(28,1)が「使用図書」でない→講義回数16（ICT活用総合実習）
        # 単位数6→講義回数15（企業内実習III：事業創造）

        tani_su = in_sheet.cell(row=6, column=5).value
        tosyo_cell = in_sheet.cell(row=21, column=1).value
        if tani_su == 1 and tosyo_cell == "【使用図書】":
            template = 'template_8'
            map_index = 0
        elif tani_su == 1:
            template = 'template_15'
            map_index = 1
        elif tani_su == 2 and tosyo_cell == "【使用図書】":
            template = 'template_8'
            map_index = 0
        elif tani_su == 2 and in_sheet.cell(row=43, column=1).value == "【使用図書】":
            template = 'template_30'
            map_index = 5
        elif tani_su == 2 :
            template = 'template_15'
            map_index = 1
        elif tani_su == 5 and in_sheet.cell(row=25, column=1).value == "【使用図書】" :
            template = 'template_12'
            map_index = 6
        elif tani_su == 5 :
            template = 'template_14'
            map_index = 2
        elif tani_su == 15 :
            template = 'template_26'
            map_index = 3
        elif tani_su == 4 and in_sheet.cell(row=28, column=1).value == "【使用図書】":
            template = 'template_15'
            map_index = 1
        elif tani_su == 4 and in_sheet.cell(row=25, column=1).value == "【使用図書】":
            template = 'template_12'
            map_index = 6
        elif tani_su == 4:
            template = 'template_16'
            map_index = 4
        elif tani_su == 6:
            template = 'template_15'
            map_index = 1
        else :
            print("error: そんな単位数はないはず: シート名 " + in_sheet.title)
            exit()

        # 出力テンプレートシート
        temp_sheet = out_wb[template]
        # テンプレートシートをコピーしてリネーム
        out_sheet = out_wb.copy_worksheet(temp_sheet)
        out_sheet.title = kamoku

        in_map = in_map_list[map_index]
        out_map = out_map_list[map_index]

        # オフセットは，「授業科目」が左上(1,1)の位置からいくつずれているか（ずれなしは(0,0)）
        # セルを読む：入力
        in_cells = read_cells(in_sheet, in_map, (1,0))

        # セルに代入
        # 新
        write_cells(out_sheet, in_cells, out_map, (2,1))
        # 旧
        write_cells(out_sheet, in_cells, out_map, (2,8))

    # ファイル保存（ここでファイルは新規作成される）
    out_wb.save(out_filename)

# 入出力ファイル
in_filename = "./input/シラバス情報（修正版）.xlsx"
out_filename = "./output/新旧対応表（情報）.xlsx"
in_filename2 = "./input/シラバス事業創造（修正版）.xlsx"
out_filename2 = "./output/新旧対応表（事業創造）.xlsx"

# Excelの中での読み書きのセル位置情報
# 授業科目の文字の入ったセルの位置を(1,1)とする
# 講義回数によって変化するから生成する

# in_mapの生成関数
def create_in_map(kougi_kaisu):
    map = [
        # 担当教員名〜時間数
        (1,3), (4,1), (3,5), (4,5), (5,5), (3,7), (4,7), (5, 7),
        # 概要，学習目標
        (7,1), (9,1),
    ]
    # 単元〜学習方法
    for i in range(kougi_kaisu) :
        map.append((12+i, 1))
        map.append((12+i, 2))
        map.append((12+i, 7))
        map.append((12+i, 8))
    # 残りを追加
    map.extend( [
        # 使用図書〜発行年
        (kougi_kaisu+13,3), (kougi_kaisu+13,5), (kougi_kaisu+13,7), (kougi_kaisu+13,9),
        (kougi_kaisu+14,3), (kougi_kaisu+14,5), (kougi_kaisu+14,7), (kougi_kaisu+14,9),
        (kougi_kaisu+15,3), (kougi_kaisu+15,5), (kougi_kaisu+15,7), (kougi_kaisu+15,9),
        # 準備学修
        (kougi_kaisu+16,3),
        # 評価方法・履修上の留意点
        (kougi_kaisu+18,1), (kougi_kaisu+18,4)
    ])
    return map

# out_mapの生成関数
def create_out_map(kougi_kaisu):
    map = [
        # 担当教員名〜時間数
        (1,2), (2,2), (3,2), (4,2), (5,2), (6,2), (7,2), (8,2),
        # 概要，学習目標
        (9,2), (10,2),
    ]
    # 単元〜学習方法
    for i in range(kougi_kaisu) :
        map.append((12+i, 1))
        map.append((12+i, 2))
        map.append((12+i, 3))
        map.append((12+i, 4))
    # 残りを追加
    map.extend( [
        # 使用図書〜発行年
        (kougi_kaisu+13,2), (kougi_kaisu+13,3), (kougi_kaisu+13,4), (kougi_kaisu+13,5),
        (kougi_kaisu+14,2), (kougi_kaisu+14,3), (kougi_kaisu+14,4), (kougi_kaisu+14,5),
        (kougi_kaisu+15,2), (kougi_kaisu+15,3), (kougi_kaisu+15,4), (kougi_kaisu+15,5),
        # 準備学修
        (kougi_kaisu+16,2),
        # 評価方法・履修上の留意点
        (kougi_kaisu+17,2), (kougi_kaisu+18,2)
        ])
    return map

# mapの生成
in_map_list = []
in_map_list.append(create_in_map(8))
in_map_list.append(create_in_map(15))
in_map_list.append(create_in_map(14))
in_map_list.append(create_in_map(26))
in_map_list.append(create_in_map(16))
in_map_list.append(create_in_map(30))
in_map_list.append(create_in_map(12))
out_map_list = []
out_map_list.append(create_out_map(8))
out_map_list.append(create_out_map(15))
out_map_list.append(create_out_map(14))
out_map_list.append(create_out_map(26))
out_map_list.append(create_out_map(16))
out_map_list.append(create_out_map(30))
out_map_list.append(create_out_map(12))

syllabus(in_filename, out_filename, in_map_list, out_map_list)
syllabus(in_filename2, out_filename2, in_map_list, out_map_list)


